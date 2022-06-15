import openpyxl
import jinja2


def excel_test_1(path):
    print(path)

    file_obj = openpyxl.load_workbook(path)

    sheet_obj = file_obj.active

    cell_obj = sheet_obj.cell(row=1, column=1)
    print(cell_obj.value)


# ============================================================

def excel_test_2(path):
    file_obj = openpyxl.load_workbook(path)
    sheet_obj = file_obj.active
    row = sheet_obj.max_row
    column = sheet_obj.max_column

    print(row, column)

    # für die FOR-Schleife muss man den Index anpassen, weil "pyxl" mit INDEX 1 arbeitet
    for i in range(1, row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        print(cell_obj.value)


# ============================================================

def excel_test_3(sheet_object_instance):
    row = sheet_object_instance.max_row
    column = sheet_object_instance.max_column
    for i in range(2, row + 1):
        for j in range(2, column + 1):
            cell_obj = sheet_object_instance.cell(row=i, column=j)
            print(cell_obj.value)


# ============================================================

# this function creates files according to a jinja2 Template
# the number of files is defined by the number of devices in the Exel file
def excel_to_dict(sheet_object_instance):
    row = sheet_object_instance.max_row
    column = sheet_object_instance.max_column

    # create empty list
    column_keys = []
    for i in range(1, column + 1):
        cell_obj = sheet_object_instance.cell(row=1, column=i)
        column_keys.append(cell_obj.value)

    # create empty dict
    devices = {}

    # open jinja template
    with open(f"templates/test_template_2.j2") as openedJinja2File:
        template_data = openedJinja2File.read()
        print('Template File opened successfully')
    template = jinja2.Template(template_data)

    for i in range(2, row + 1):
        # create new element in dict
        devices[i-1] = {}

        for j in range(2, column + 1):
            cell_obj = sheet_object_instance.cell(row=i, column=j)
            # add current attribute with key to the current element
            devices[i-1][column_keys[j-1]] = cell_obj.value

        new_file_name = devices[i-1].get("Geraetename")

        # create a file with the template for each iteration
        with open(f"output/{new_file_name}.txt", 'w') as new_file:
            new_file.write(template.render(devices=devices[i-1]))
            new_file.close()

    openedJinja2File.close()

# ============================================================


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    path = "Mappe1.xlsx"



    # excel_test_1(path)
    # excel_test_2(path)

    file_obj = openpyxl.load_workbook(path)
    sheet_obj = file_obj.active

    # excel_test_3(sheet_obj)
    excel_to_dict(sheet_obj)
